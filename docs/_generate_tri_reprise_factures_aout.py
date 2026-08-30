"""Génère docs/Tri-reprise-factures-aout.xlsx pour le tri de reprise."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = ROOT / "database" / "seeders" / "data" / "checklist_reprise_aout_factures.json"
OUTPUT = ROOT / "docs" / "Tri-reprise-factures-aout.xlsx"


def main() -> None:
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Tri reprise"

    headers = [
        "N° registre",
        "Date réception",
        "Fournisseur",
        "Objet (brut checklist)",
        "Montant",
        "Téléphone",
        "Payé ? (checklist)",
        "État réel",
        "Étape physique",
        "Mode paiement",
        "Instructions DG (résumé)",
        "Preuve BPA",
        "Réf. pièce",
        "Banque",
        "Date BPA",
        "Montant payé",
        "Date paiement / décharge",
        "Notifier fournisseur ?",
        "Action COSUD",
        "Priorité reprise",
        "Notes",
    ]

    header_fill = PatternFill("solid", fgColor="00695C")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="B0BEC5"),
        right=Side(style="thin", color="B0BEC5"),
        top=Side(style="thin", color="B0BEC5"),
        bottom=Side(style="thin", color="B0BEC5"),
    )
    alt_fill = PatternFill("solid", fgColor="E0F2F1")
    wrap = Alignment(wrap_text=True, vertical="center")

    for col, title in enumerate(headers, 1):
        cell = ws.cell(1, col, title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "B2"
    last_row = 1 + len(data)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    for i, row in enumerate(data):
        r = i + 2
        values = [
            row.get("numero_registre_complet"),
            row.get("date_reception"),
            row.get("expediteur_libelle"),
            row.get("objet_brut"),
            row.get("montant"),
            row.get("telephone"),
            row.get("paye"),
            "inconnu",
            "—",
            "?",
            "",
            "non",
            "",
            "",
            "",
            "",
            "",
            "non",
            "a_completer",
            3,
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(r, col, val)
            cell.border = thin
            cell.alignment = wrap
            if i % 2 == 1:
                cell.fill = alt_fill
        ws.row_dimensions[r].height = 28

    validations = [
        ("H", '"a_instruire,en_cours,paye,inconnu"'),
        ("I", '"—,bpa,ac_etablit,dg_signe,decharge,cloture"'),
        ("J", '"cheque,ov,?"'),
        ("L", '"oui,non,partiel"'),
        ("R", '"oui,non"'),
        ("S", '"attendre_dg,saisir_bpa,avancer_circuit,regularisation_payee,a_completer"'),
        ("T", '"1,2,3"'),
    ]
    for col, formula in validations:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.add(f"{col}2:{col}{last_row}")
        ws.add_data_validation(dv)

    widths = {
        "A": 12,
        "B": 13,
        "C": 28,
        "D": 45,
        "E": 12,
        "F": 18,
        "G": 12,
        "H": 14,
        "I": 14,
        "J": 12,
        "K": 28,
        "L": 12,
        "M": 14,
        "N": 12,
        "O": 12,
        "P": 12,
        "Q": 16,
        "R": 14,
        "S": 18,
        "T": 10,
        "U": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    leg = wb.create_sheet("Légende", 1)
    leg["A1"] = "Tri reprise factures août 2026 — mode d’emploi"
    leg["A1"].font = Font(bold=True, size=14, color="00695C", name="Calibri")
    leg.merge_cells("A1:C1")

    legende_rows = [
        ("Colonne", "Valeurs", "Signification"),
        ("État réel", "a_instruire", "Pas encore de BPA / à instruire"),
        ("État réel", "en_cours", "BPA fait, paiement pas terminé"),
        ("État réel", "paye", "Déjà payé (décharge / accusé banque)"),
        ("État réel", "inconnu", "Preuves pas encore rassemblées (défaut)"),
        ("Étape physique", "bpa", "Bon pour accord DG fait"),
        ("Étape physique", "ac_etablit", "AC a établi chèque / OV"),
        ("Étape physique", "dg_signe", "DG a signé chèque / OV"),
        ("Étape physique", "decharge", "En attente décharge / accusé banque"),
        ("Étape physique", "cloture", "Circuit terminé côté papier"),
        ("Mode paiement", "cheque / ov / ?", "Obligatoire dès que Preuve BPA = oui"),
        ("Preuve BPA", "oui / non / partiel", "Document instructions DG en main ?"),
        ("Notifier fournisseur ?", "non (reprise historique)", "oui seulement si dossier encore ouvert à notifier"),
        ("Action COSUD", "attendre_dg", "Laisser en attente instructions DG"),
        ("Action COSUD", "saisir_bpa", "Saisir BPA + mode dans COSUD"),
        ("Action COSUD", "avancer_circuit", "Avancer jusqu’à l’étape physique"),
        ("Action COSUD", "regularisation_payee", "Facture déjà payée — régularisation / clôture"),
        ("Action COSUD", "a_completer", "Manque de preuves — ne pas toucher COSUD"),
        ("Priorité reprise", "1", "Preuves OK — prêt à saisir"),
        ("Priorité reprise", "2", "Partiel — bientôt"),
        ("Priorité reprise", "3", "En attente de documents (défaut)"),
    ]

    leg_header_fill = PatternFill("solid", fgColor="00695C")
    for i, row in enumerate(legende_rows, 3):
        for c, val in enumerate(row, 1):
            cell = leg.cell(i, c, val)
            cell.border = thin
            if i == 3:
                cell.fill = leg_header_fill
                cell.font = Font(bold=True, color="FFFFFF")
            elif c == 1:
                cell.font = Font(bold=True)

    leg.column_dimensions["A"].width = 22
    leg.column_dimensions["B"].width = 36
    leg.column_dimensions["C"].width = 55

    leg["A26"] = "Ordre de travail"
    leg["A26"].font = Font(bold=True, size=12, color="00695C")
    leg["A27"] = "1. Remplir d’abord Preuve BPA + État réel (même si le reste est vide)."
    leg["A28"] = "2. Dès Preuve BPA = oui et mode connu → Priorité 1 + Action COSUD adaptée."
    leg["A29"] = "3. Ne toucher COSUD que pour les priorités 1."
    leg["A30"] = "4. Les lignes inconnu / preuve non restent en attente jusqu’aux documents DG."
    leg["A31"] = "5. Pour les dossiers déjà payés : Notifier fournisseur ? = non."

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"OK {OUTPUT} ({len(data)} lignes)")


if __name__ == "__main__":
    main()
